import wandb
import openpyxl
from argparse import Namespace
from collections import defaultdict

api = wandb.Api(timeout=90)
ds = ['IIIT_WVU', 'NotreDame', 'Clarkson']

params = {d: {dd: {'bs': 0, 'lambda': 0.0} for dd in ds} for d in ds}

params['IIIT_WVU']['IIIT_WVU']  = {'bs': 64 , 'lambda': 0.25}
params['IIIT_WVU']['NotreDame'] = {'bs': 128, 'lambda': 0.75}
params['IIIT_WVU']['Clarkson']  = {'bs': 512, 'lambda': 1.0}

params['NotreDame']['IIIT_WVU']  = {'bs': 512, 'lambda': 0.50}
params['NotreDame']['NotreDame'] = {'bs': 128, 'lambda': 0.25}
params['NotreDame']['Clarkson']  = {'bs': 512, 'lambda': 0.25}

params['Clarkson']['IIIT_WVU']  = {'bs': 256, 'lambda': 2.0}
params['Clarkson']['NotreDame'] = {'bs': 256, 'lambda': 0.25}
params['Clarkson']['Clarkson']  = {'bs': 128, 'lambda': 2.0}

runs = {d: defaultdict(list) for d in ds}

for train in ds:
    for run in api.runs(f"{train}"):
        runs[train][run.config['CLASSIFIER']['model']['type'].replace('torchvision.models.', "")].append(run)

models = sorted(list(set(runs['NotreDame'].keys()).union(set(runs['IIIT_WVU'].keys()), set(runs['Clarkson'].keys()))))

results = {model:{train:{test:Namespace(ACER=0, APCER=0, BPCER=0, invoked=False, run=None) for test in ds} for train in ds} for model in models}

for train in ds:
    for model in runs[train]:
        for run in runs[train][model]:
            batch_size = run.config['CLASSIFIER']['equiv_batch_size']
            lambda_nce = run.config['CUT']['lambda_NCE']
            for test, args in params[train].items():
                if args['bs'] == batch_size and args['lambda'] == lambda_nce:
                    if not results[model][train][test].invoked:
                        results[model][train][test].run = run
                        results[model][train][test].invoked = True
                        results[model][train][test].ACER = run.summary[f"acer/{test}/TEST"]['min']
#                     else:
#                         results[model][train][test].ACER = min(results[model][train][test], run.summary[f"acer/{test}/TEST"]['min'])


for model in models:
    for train in ds:
        for test in ds:
            if results[model][train][test].invoked:
                run = results[model][train][test].run
                for info in run.scan_history(keys=[f"apcer/{test}/TEST", f"bpcer/{test}/TEST", f"acer/{test}/TEST"]):
                    apcer = info[f"apcer/{test}/TEST"]
                    acer = info[f"acer/{test}/TEST"]
                    bpcer = info[f"bpcer/{test}/TEST"]
                    if abs(acer-results[model][train][test].ACER)/acer < 1.0e-5:
                        results[model][train][test].APCER = apcer
                        results[model][train][test].BPCER = bpcer
                
                results[model][train][test].ACER = "{:.02f}".format(results[model][train][test].ACER*100)
                results[model][train][test].APCER = "{:.02f}".format(results[model][train][test].APCER*100)
                results[model][train][test].BPCER = "{:.02f}".format(results[model][train][test].BPCER*100)
            else:
                results[model][train][test].ACER = "-"
                results[model][train][test].APCER = "-"
                results[model][train][test].BPCER = "-"

model_name = 'PBS'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['IIIT_WVU']['NotreDame'].ACER = 16.86
results[model_name]['IIIT_WVU']['Clarkson'].ACER = 47.17
results[model_name]['NotreDame']['IIIT_WVU'].ACER = 17.49
results[model_name]['NotreDame']['Clarkson'].ACER = 45.31
results[model_name]['Clarkson']['IIIT_WVU'].ACER = 42.28
results[model_name]['Clarkson']['NotreDame'].ACER = 32.42
models.append(model_name)

model_name = 'A-PBS'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['IIIT_WVU']['NotreDame'].ACER = 27.61
results[model_name]['IIIT_WVU']['Clarkson'].ACER = 21.99
results[model_name]['NotreDame']['IIIT_WVU'].ACER = 9.49
results[model_name]['NotreDame']['Clarkson'].ACER = 22.46
results[model_name]['Clarkson']['IIIT_WVU'].ACER = 34.17
results[model_name]['Clarkson']['NotreDame'].ACER = 23.08
models.append(model_name)

model_name = 'FAM+FMM'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['IIIT_WVU']['NotreDame'].ACER = 5.81
results[model_name]['IIIT_WVU']['Clarkson'].ACER = 26.03
results[model_name]['NotreDame']['IIIT_WVU'].ACER = 15.07
results[model_name]['NotreDame']['Clarkson'].ACER = 10.51
results[model_name]['Clarkson']['IIIT_WVU'].ACER = 22.06
results[model_name]['Clarkson']['NotreDame'].ACER = 20.92
models.append(model_name)

model_name = 'CASIA'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['Clarkson']['Clarkson'].ACER = 7.10
results[model_name]['NotreDame']['NotreDame'].ACER = 4.03
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 16.70
models.append(model_name)

model_name = 'SpoofNet'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['Clarkson']['Clarkson'].ACER = 16.50
results[model_name]['NotreDame']['NotreDame'].ACER = 9.50
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 18.62
models.append(model_name)

model_name = 'D-NetPAD'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['Clarkson']['Clarkson'].ACER = 3.36
results[model_name]['NotreDame']['NotreDame'].ACER = 6.81
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 23.27
models.append(model_name)

model_name = 'MSA'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['NotreDame']['NotreDame'].ACER = 6.23
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 11.13
models.append(model_name)

model_name = 'PBS'
# results[model_name] = {train:{test:"-" for test in ds} for train in ds}
results[model_name]['Clarkson']['Clarkson'].ACER = 4.48
results[model_name]['NotreDame']['NotreDame'].ACER = 4.97
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 7.01
# models.append(model_name)

model_name = 'A-PBS'
# results[model_name] = {train:{test:"-" for test in ds} for train in ds}
results[model_name]['Clarkson']['Clarkson'].ACER = 3.48
results[model_name]['NotreDame']['NotreDame'].ACER = 3.94
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 6.50
# models.append(model_name)

model_name = 'FAM'
results[model_name] = {train: {test: Namespace(ACER="-", APCER="-", BPCER="-") for test in ds} for train in ds}
results[model_name]['Clarkson']['Clarkson'].ACER = 3.45
results[model_name]['NotreDame']['NotreDame'].ACER = 4.03
results[model_name]['IIIT_WVU']['IIIT_WVU'].ACER = 6.84
models.append(model_name)

wb = openpyxl.Workbook()

sheet = wb.create_sheet('ACER')
sheet.append("Train,IIIT_WVU,,,NotreDame,,,Clarkson,,,\n".strip().split(','))
sheet.append("Test,IIIT_WVU,NotreDame,Clarkson,IIIT_WVU,NotreDame,Clarkson,IIIT_WVU,NotreDame,Clarkson\n".strip().split(','))
for model in models:
    msg = f"{model},"
    for train in ds:
        for test in ds:
            msg += f"{results[model][train][test].ACER},"
    sheet.append((msg + "\n").strip().split(','))
    
sheet = wb.create_sheet('APCER')
sheet.append("Train,IIIT_WVU,,,NotreDame,,,Clarkson,,,\n".strip().split(','))
sheet.append("Test,IIIT_WVU,NotreDame,Clarkson,IIIT_WVU,NotreDame,Clarkson,IIIT_WVU,NotreDame,Clarkson\n".strip().split(','))
for model in models:
    msg = f"{model},"
    for train in ds:
        for test in ds:
            msg += f"{results[model][train][test].APCER},"
    sheet.append((msg + "\n").strip().split(','))

sheet = wb.create_sheet('BPCER')
sheet.append("Train,IIIT_WVU,,,NotreDame,,,Clarkson,,,\n".strip().split(','))
sheet.append("Test,IIIT_WVU,NotreDame,Clarkson,IIIT_WVU,NotreDame,Clarkson,IIIT_WVU,NotreDame,Clarkson\n".strip().split(','))
for model in models:
    msg = f"{model},"
    for train in ds:
        for test in ds:
            msg += f"{results[model][train][test].BPCER},"
    sheet.append((msg + "\n").strip().split(','))
    
wb.save('LivDet-iris-2017.xlsx')
